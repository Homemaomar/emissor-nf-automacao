from pathlib import Path
from datetime import datetime
import unicodedata

import pandas as pd
from openpyxl import load_workbook


class PlanilhaNotasRepository:

    def __init__(self, caminho_planilha):
        self.caminho_planilha = Path(caminho_planilha)

        if not self.caminho_planilha.exists():
            raise FileNotFoundError(
                f"Planilha não encontrada: {self.caminho_planilha}"
            )

        self.df = self._carregar_planilha()

    def _normalizar_coluna(self, coluna):
        texto = str(coluna or "").strip().upper()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
        return " ".join(texto.split())

    def _valor_coluna_linha(self, row, coluna):
        if not coluna:
            return ""

        valor = row.get(coluna, "")
        texto = str(valor or "").strip()
        if not texto or self._normalizar_coluna(texto) == coluna:
            return ""

        return valor

    def _carregar_planilha(self):
        xls = pd.ExcelFile(self.caminho_planilha, engine="openpyxl")
        abas_normalizadas = [a.upper().strip() for a in xls.sheet_names]

        if "NOTAS" in abas_normalizadas:
            nome_aba_real = xls.sheet_names[abas_normalizadas.index("NOTAS")]
        else:
            raise Exception("Aba 'NOTAS' não encontrada")

        df = pd.read_excel(
            self.caminho_planilha,
            sheet_name=nome_aba_real,
            engine="openpyxl",
        )
        xls.close()

        print("COLUNAS ORIGINAIS:", df.columns.tolist())
        df.columns = [self._normalizar_coluna(col) for col in df.columns]
        print("COLUNAS NORMALIZADAS:", df.columns.tolist())

        colunas_necessarias = ["STATUS", "ITEM"]
        for col in colunas_necessarias:
            if col not in df.columns:
                raise Exception(
                    f"Coluna obrigatória '{col}' não encontrada.\n"
                    f"Colunas disponíveis: {df.columns.tolist()}"
                )

        df["STATUS"] = df["STATUS"].astype(str)
        df["ITEM"] = df["ITEM"].astype(str)

        return df

    def _coluna_cliente(self, df):
        if "SECRETARIA" in df.columns:
            return "SECRETARIA"
        if "CLIENTE" in df.columns:
            return "CLIENTE"
        if "CLIENTE.1" in df.columns:
            return "CLIENTE.1"
        return None

    def _coluna_descricao(self, df):
        if "DESCRICAO" in df.columns:
            return "DESCRICAO"
        return None

    def _coluna_especie(self, df):
        if "ESPECIE" in df.columns:
            return "ESPECIE"
        return None

    def obter_dados_item(self, item_id):
        """
        Recarrega a planilha e retorna os dados atualizados de um item específico.
        """
        df = self._carregar_planilha()
        linha = df[df["ITEM"] == str(item_id)]

        if linha.empty:
            raise Exception(f"Item {item_id} não encontrado na planilha.")

        return linha.iloc[0].to_dict()

    def listar_notas_pendentes(self, cliente=None, especie=None, itens=None):
        df = self.df.copy()
        coluna_cliente = self._coluna_cliente(df)
        coluna_descricao = self._coluna_descricao(df)
        coluna_especie = self._coluna_especie(df)

        print("COLUNAS DO DF:", df.columns.tolist())
        print("COLUNA CLIENTE:", coluna_cliente)
        print("COLUNA DESCRICAO:", coluna_descricao)
        print("COLUNA ESPECIE:", coluna_especie)

        df = df[df["STATUS"].str.upper() == "PENDENTE"]

        if cliente and cliente != "Todos":
            if coluna_cliente:
                df = df[df[coluna_cliente] == cliente]
            else:
                df = df.iloc[0:0]

        if especie and especie != "Todas":
            if coluna_especie:
                df = df[df[coluna_especie] == especie]
            else:
                df = df.iloc[0:0]

        if itens:
            df = df[df["ITEM"].isin(itens)]

        df = df.reset_index()
        notas = []

        for _, row in df.iterrows():
            notas.append({
                "excel_row": int(row["index"]) + 2,
                "item": row.get("ITEM", ""),
                "cliente": self._valor_coluna_linha(row, coluna_cliente),
                "descricao": self._valor_coluna_linha(row, coluna_descricao),
                "valor": row.get("VALOR", 0),
                "ir": row.get("IR", 0),
                "iss": row.get("ISS", 0),
                "cnpj": row.get("CNPJ", ""),
                "ctn": row.get("CTN", ""),
                "nbs": row.get("NBS", ""),
                "email": row.get("EMAIL", ""),
                "especie": self._valor_coluna_linha(row, coluna_especie),
            })

        return notas

    def atualizar_resultado_emissao(
        self,
        excel_row,
        status,
        usuario,
        numero_nfse="",
        caminho_xml="",
        caminho_pdf="",
        erro="",
    ):
        wb = load_workbook(self.caminho_planilha, keep_vba=True)
        abas_normalizadas = {
            self._normalizar_coluna(nome_aba): nome_aba for nome_aba in wb.sheetnames
        }
        nome_aba = abas_normalizadas.get("NOTAS")
        if not nome_aba:
            wb.close()
            raise Exception("Aba 'NOTAS' não encontrada para atualizar a emissão.")

        ws = wb[nome_aba]

        headers = {}
        for col in range(1, ws.max_column + 1):
            nome = ws.cell(row=1, column=col).value
            if nome:
                headers[self._normalizar_coluna(nome)] = col

        def set_val(coluna, valor):
            coluna_normalizada = self._normalizar_coluna(coluna)
            if coluna_normalizada not in headers:
                nova_coluna = ws.max_column + 1
                ws.cell(row=1, column=nova_coluna, value=coluna)
                headers[coluna_normalizada] = nova_coluna
            ws.cell(row=excel_row, column=headers[coluna_normalizada], value=valor)

        set_val("STATUS", status)
        set_val("EMITIDA_POR", usuario)

        if status == "EMITIDA":
            set_val("NUMERO_NFSE", numero_nfse)
            set_val("DATA_EMISSAO", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            set_val("CAMINHO_XML", caminho_xml)
            set_val("CAMINHO_PDF", caminho_pdf)

        if status == "ERRO":
            set_val("ERRO", erro)

        wb.save(self.caminho_planilha)
        wb.close()


def montar_caminho_planilha(base_notas, ano, mes, municipio):
    base = Path(base_notas)
    mes = str(mes or "").strip()
    municipio = str(municipio or "").strip()

    if not base.exists():
        raise FileNotFoundError(f"Caminho base não encontrado:\n{base}")

    pasta = base / ano / mes / municipio
    aliases_mes = {
        "03 - Marco": "03 - Março",
        "03 - Março": "03 - Marco",
    }

    if not pasta.exists() and mes in aliases_mes:
        pasta_alternativa = base / ano / aliases_mes[mes] / municipio
        if pasta_alternativa.exists():
            pasta = pasta_alternativa

    if not pasta.exists():
        raise FileNotFoundError(f"Pasta não encontrada:\n{pasta}")

    arquivo_xlsx = pasta / "notas.xlsx"
    arquivo_xlsm = pasta / "notas.xlsm"

    if arquivo_xlsm.exists():
        return str(arquivo_xlsm)

    if arquivo_xlsx.exists():
        return str(arquivo_xlsx)

    arquivos_encontrados = list(pasta.glob("*"))
    raise FileNotFoundError(
        f"Nenhuma planilha encontrada em:\n{pasta}\n\n"
        f"Arquivos encontrados:\n"
        + "\n".join(str(a.name) for a in arquivos_encontrados)
    )
