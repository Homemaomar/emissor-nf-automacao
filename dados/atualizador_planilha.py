from openpyxl import load_workbook


class AtualizadorPlanilha:
    def __init__(self, nome_aba="NOTAS"):
        self.nome_aba = nome_aba

    def atualizar_resultado(
        self,
        caminho_planilha,
        excel_row,
        status="",
        numero_nfse="",
        data_emissao="",
        caminho_xml="",
        caminho_pdf="",
        mensagem=""
    ):
        wb = load_workbook(self.caminho_planilha, keep_vba=True)
        ws = wb[self.nome_aba]

        headers = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor:
                headers[str(valor).strip().upper()] = col

        if "STATUS" in headers:
            ws.cell(row=excel_row, column=headers["STATUS"]).value = status

        if "NUMERO_NFSE" in headers:
            ws.cell(row=excel_row, column=headers["NUMERO_NFSE"]).value = numero_nfse

        if "DATA_EMISSAO" in headers:
            ws.cell(row=excel_row, column=headers["DATA_EMISSAO"]).value = data_emissao

        if "CAMINHO_XML" in headers:
            ws.cell(row=excel_row, column=headers["CAMINHO_XML"]).value = caminho_xml

        if "CAMINHO_PDF" in headers:
            ws.cell(row=excel_row, column=headers["CAMINHO_PDF"]).value = caminho_pdf

        # Se quiser, você pode criar uma coluna MENSAGEM futuramente
        wb.save(caminho_planilha)
        wb.close()