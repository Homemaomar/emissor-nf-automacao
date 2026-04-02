from openpyxl import load_workbook

from .utils_envio import agora_str


class AtualizadorStatusEnvio:
    def __init__(self, caminho_planilha: str, nome_aba: str = "NOTAS", log_callback=None):
        self.caminho_planilha = caminho_planilha
        self.nome_aba = nome_aba
        self.log = log_callback or print

    def _garantir_coluna(self, ws, nome_coluna: str):
        headers = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor:
                headers[str(valor).strip().upper()] = col

        nome_normalizado = nome_coluna.strip().upper()
        if nome_normalizado in headers:
            return headers[nome_normalizado]

        nova_coluna = ws.max_column + 1
        ws.cell(row=1, column=nova_coluna).value = nome_coluna
        return nova_coluna

    def atualizar_status_email(self, linhas_excel, status, erro="", protocolo=""):
        wb = load_workbook(self.caminho_planilha, keep_vba=True)
        abas = wb.sheetnames
        abas_normalizadas = [a.upper().strip() for a in abas]

        if self.nome_aba.upper() in abas_normalizadas:
            nome_real = abas[abas_normalizadas.index(self.nome_aba.upper())]
        else:
            raise Exception(f"Aba {self.nome_aba} não encontrada. Abas disponíveis: {abas}")

        ws = wb[nome_real]

        col_status = self._garantir_coluna(ws, "STATUS_EMAIL")
        col_erro = self._garantir_coluna(ws, "ERRO_EMAIL")
        col_data = self._garantir_coluna(ws, "DATA_ENVIO_EMAIL")
        col_protocolo = self._garantir_coluna(ws, "PROTOCOLO_EMAIL")

        for linha in linhas_excel:
            ws.cell(row=linha, column=col_status).value = status
            ws.cell(row=linha, column=col_erro).value = erro
            ws.cell(row=linha, column=col_data).value = agora_str()
            ws.cell(row=linha, column=col_protocolo).value = protocolo

        wb.save(self.caminho_planilha)
        self.log(f"📝 STATUS_EMAIL atualizado para {len(linhas_excel)} linha(s).")

    def atualizar_status_whatsapp(self, linhas_excel, status, erro="", protocolo=""):
        wb = load_workbook(self.caminho_planilha, keep_vba=True)
        abas = wb.sheetnames
        abas_normalizadas = [a.upper().strip() for a in abas]

        if self.nome_aba.upper() in abas_normalizadas:
            nome_real = abas[abas_normalizadas.index(self.nome_aba.upper())]
        else:
            raise Exception(f"Aba {self.nome_aba} não encontrada. Abas disponíveis: {abas}")

        ws = wb[nome_real]

        col_status = self._garantir_coluna(ws, "STATUS_WHATSAPP")
        col_erro = self._garantir_coluna(ws, "ERRO_WHATSAPP")
        col_data = self._garantir_coluna(ws, "DATA_ENVIO_WHATSAPP")
        col_protocolo = self._garantir_coluna(ws, "PROTOCOLO_WHATSAPP")

        for linha in linhas_excel:
            ws.cell(row=linha, column=col_status).value = status
            ws.cell(row=linha, column=col_erro).value = erro
            ws.cell(row=linha, column=col_data).value = agora_str()
            ws.cell(row=linha, column=col_protocolo).value = protocolo

        wb.save(self.caminho_planilha)
        self.log(f"📝 STATUS_WHATSAPP atualizado para {len(linhas_excel)} linha(s).")